import json
import os
from src.auth import perform_login
from src.processor import process_and_save_data, append_mobialert_to_excel
from src.email_sender import send_report_email
from src import config
from src.utils import setup_logger
from test import get_mobialert_failures

logger = setup_logger("main")

def main():
    logger.info("Starting Scraper Job")
    
    # 1. Authenticate
    session, csrf_token = perform_login()
    
    if not session or not csrf_token:
        logger.critical("Authentication failed. Aborting.")
        return

    # 2. Fetch Analytical Portal Data
    logger.info(f"GET {config.TARGET_API}")
    api_headers = {
        'X-CSRF-TOKEN': csrf_token, 
        'X-Requested-With': 'XMLHttpRequest'
    }
    
    try:
        history_response = session.get(config.TARGET_API, headers=api_headers, verify=False)
        
        if history_response.status_code == 200:
            logger.info("Data fetched successfully")
            
            raw_data = history_response.json()

            # 3. Process Analytical Portal Data → Excel
            generated_file_path = process_and_save_data(raw_data)

            # 4. Fetch MobiAlert failures from SAP HANA and append to the same Excel
            logger.info("Querying MobiAlert failures from SAP HANA...")
            mobialert_rows = get_mobialert_failures()

            if mobialert_rows:
                logger.info(f"Found {len(mobialert_rows)} MobiAlert failure(s).")
            else:
                logger.info("No MobiAlert failures found for yesterday.")

            if generated_file_path and os.path.exists(generated_file_path):
                # Append MobiAlert section (even if empty list – function handles it gracefully)
                generated_file_path = append_mobialert_to_excel(
                    generated_file_path, mobialert_rows
                )
            elif mobialert_rows:
                # Analytical portal is healthy but there ARE MobiAlert failures –
                # create an Excel with only the MobiAlert section.
                logger.info(
                    "No portal failures found, but MobiAlert has failures. "
                    "Creating Excel with MobiAlert section only."
                )
                # Build a minimal stub Excel so append_mobialert_to_excel has somewhere to write
                import openpyxl
                from datetime import datetime
                from openpyxl.styles import Font, Alignment
                from src import config as cfg

                timestamp  = datetime.now().strftime("%d-%m-%Y %I-%M-%p")
                output_dir = getattr(cfg, 'OUTPUT_DIRECTORY', '.')
                os.makedirs(output_dir, exist_ok=True)
                prefix     = getattr(cfg, 'REPORT_FILENAME_PREFIX', 'failed_reports')
                stub_path  = os.path.join(output_dir, f"{prefix}_{timestamp}.xlsx")

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Failed Reports"
                ws.merge_cells('A1:F1')
                c = ws['A1']
                c.value     = f"{datetime.now().year}'s Schedule Not Sent by Email"
                c.font      = Font(size=14, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")

                ws.merge_cells('A2:F2')
                ws['A2'].value = "ℹ️ Analytical Portal: No failures found. See MobiAlert section below."
                wb.save(stub_path)

                generated_file_path = append_mobialert_to_excel(stub_path, mobialert_rows)

            # # 5. Send Email (only if a file was generated)
            # if generated_file_path and os.path.exists(generated_file_path):
            #     logger.info(f"Report generated at {generated_file_path}. Sending email...")
                
            #     email_status = send_report_email(generated_file_path)
                
            #     if email_status:
            #         logger.info("Email sent successfully.")
            #     else:
            #         logger.error("Email sending failed.")
            # else:
            #     logger.info("No failures found in either system – all healthy. Skipping email.")

        else:
            logger.error(f"Failed to fetch history. Status: {history_response.status_code}")
            
    except Exception as e:
        logger.error(f"Main loop exception: {e}")

if __name__ == "__main__":
    main()
