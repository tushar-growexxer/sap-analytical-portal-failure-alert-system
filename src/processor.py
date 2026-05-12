import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter  # <--- Added this import
from datetime import datetime, timedelta
import os
from src.utils import setup_logger
from src import config

logger = setup_logger("processor")

def process_and_save_data(json_data):
    """
    1. Parses history and identifies the latest status based on reportId_scheduleName.
    2. Filters for the CURRENT YEAR only.
    3. Filters for problematic statuses in the most recent execution.
    4. Saves results to Excel with Title and Legend.
    """
    logger.info("Processing History Data...")
    current_year = datetime.now().year
    
    try:
        # Extract history list (handling different structures)
        if 'response' in json_data and 'data' in json_data['response']:
            data_source = json_data['response']['data']
            if isinstance(data_source, dict) and 'histories' in data_source:
                history_list = data_source['histories']
            elif isinstance(data_source, list):
                history_list = data_source
            else:
                history_list = next((v for v in data_source.values() if isinstance(v, list)), [])
        else:
            history_list = json_data
            
        if not history_list:
            logger.warning("No history data found to process.")
            return

        logger.info(f"Found {len(history_list)} raw records.")

        # --- STEP 1: Map Latest Data by composite key ---
        latest_map = {}

        for item in history_list:
            report_id = item.get('reportId')
            schedule_name = item.get('scheduleName')
            exec_time_str = item.get('executeTime')
            
            if not report_id or not schedule_name or not exec_time_str:
                continue

            try:
                current_exec_time = datetime.strptime(exec_time_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue

            # Filter: Current Year Only
            if current_exec_time.year != current_year:
                continue

            # Composite Key: reportId + scheduleName
            composite_key = f"{report_id}_{schedule_name}"
            
            # Logic: Update only if newer
            if composite_key not in latest_map:
                latest_map[composite_key] = {**item, 'dt_obj': current_exec_time}
            else:
                if current_exec_time > latest_map[composite_key]['dt_obj']:
                    latest_map[composite_key] = {**item, 'dt_obj': current_exec_time}

        # --- STEP 2: Filter for Failures ---
        problematic_records = []

        for record in latest_map.values():
            status = record.get('executeStatus', '').lower()
            msg = record.get('executeMsg', {}).get('msg', '')
            
            is_failure = False
            if status in ['failed', 'n/a']:
                is_failure = True
            elif status == 'succeeded' and any(err in msg.lower() for err in ['failed to send', 'login attempts']):
                is_failure = True
            
            if is_failure:
                problematic_records.append(record)

        if not problematic_records:
            logger.info(f"All latest report schedules for {current_year} are healthy.")
            return None

        # Sort by time descending
        problematic_records.sort(key=lambda x: x['dt_obj'], reverse=True)

        return save_to_excel(problematic_records, current_year)

    except Exception as e:
        logger.error(f"Error processing data: {e}")
        return None

def save_to_excel(report_list, year):
    timestamp = datetime.now().strftime("%d-%m-%Y %I-%M-%p")
    output_dir = getattr(config, 'OUTPUT_DIRECTORY', '.') 
    os.makedirs(output_dir, exist_ok=True)

    prefix = getattr(config, 'REPORT_FILENAME_PREFIX', 'failed_reports')
    output_file = os.path.join(output_dir, f"{prefix}_{timestamp}.xlsx")
    
    # Dates for Highlighting
    today_date = datetime.now().date()
    yesterday_date = today_date - timedelta(days=1)
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Reports"
        
        # --- STYLES ---
        title_font = Font(size=14, bold=True, color="000000")
        legend_font = Font(size=10, italic=True, color="555555")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Dark Red
        
        # Light Yellow for Today/Yesterday rows
        highlight_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") 
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # --- ROW 1: MAIN TITLE ---
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"{year}'s Schedule Not Sent by Email"
        title_cell.font = title_font
        title_cell.alignment = center_align

        # --- ROW 2: LEGEND ---
        ws.merge_cells('A2:F2')
        legend_cell = ws['A2']
        legend_cell.value = "⚠️ Legend: Rows highlighted in YELLOW indicate failures occurred Today or Yesterday."
        legend_cell.font = legend_font
        legend_cell.alignment = left_align
        legend_cell.fill = highlight_fill

        # --- ROW 3: HEADERS ---
        headers = ['RCRI Code', 'Report Name', 'Schedule Name', 'Execute Time', 'Status', 'Message']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # --- ROW 4+: DATA ---
        for row_idx, item in enumerate(report_list, 4): # Start at Row 4
            exec_time_str = item.get('executeTime')
            
            should_highlight = False
            try:
                row_date = datetime.strptime(exec_time_str, "%Y-%m-%d %H:%M:%S").date()
                if row_date == today_date or row_date == yesterday_date:
                    should_highlight = True
            except:
                pass 

            # Helper to write
            def write_cell(col, val):
                c = ws.cell(row=row_idx, column=col, value=val)
                if should_highlight:
                    c.fill = highlight_fill
                c.border = thin_border
                return c

            write_cell(1, item.get('reportId'))
            write_cell(2, item.get('reportName'))
            write_cell(3, item.get('scheduleName'))
            write_cell(4, exec_time_str)
            write_cell(5, item.get('executeStatus'))
            write_cell(6, item.get('executeMsg', {}).get('msg'))

        # --- SAFE COLUMN WIDTH ADJUSTMENT ---
        # Instead of ws.columns (which breaks on merged cells), iterate by index 1 to 6
        for col_idx in range(1, 7):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            
            # Check Title Row (approx length divided by columns)
            if col_idx == 1: 
                max_len = 20 # Minimum width for first column
            
            # Iterate through data rows only (skip merged title/legend)
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
        
        wb.save(output_file)
        logger.info(f"Report generated with legend: {output_file}")

        return output_file
        
    except Exception as e:
        logger.error(f"Excel generation failed: {e}")

def append_mobialert_to_excel(excel_path, mobialert_rows):
    """
    Appends MobiAlert failure rows to an existing Excel file produced by
    save_to_excel().  Adds a section header (merged, dark-blue) followed by
    column headers and data rows.

    Parameters
    ----------
    excel_path     : str  – path to the xlsx file to modify
    mobialert_rows : list – list of dicts with keys CONFIGID, CONFIGNAME,
                            EMAILMESSAGE (as returned by test.get_mobialert_failures)

    Returns the same excel_path on success, None on failure.
    """
    if not mobialert_rows:
        logger.info("No MobiAlert failures to append – skipping section.")
        return excel_path

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # ── Styles ────────────────────────────────────────────────────────────
        section_font  = Font(size=13, bold=True, color="FFFFFF")
        section_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79",
                                    fill_type="solid")   # dark navy-blue
        header_font   = Font(bold=True, color="FFFFFF")
        header_fill   = PatternFill(start_color="2E75B6", end_color="2E75B6",
                                    fill_type="solid")   # medium blue
        center_align  = Alignment(horizontal="center", vertical="center")
        left_align    = Alignment(horizontal="left",   vertical="center")
        thin_border   = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        # ── Find the next empty row (leave one blank row as spacer) ───────────
        next_row = ws.max_row + 2          # +1 blank spacer, +1 for actual start

        # ── Section title row ─────────────────────────────────────────────────
        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row,   end_column=3
        )
        title_cell = ws.cell(row=next_row, column=1,
                             value="MobiAlert – Failed to Send (Yesterday & Today)")
        title_cell.font      = section_font
        title_cell.fill      = section_fill
        title_cell.alignment = center_align
        next_row += 1

        # ── Column headers ────────────────────────────────────────────────────
        mob_headers = ["Config ID", "Config Name", "Email Message"]
        for col_idx, header in enumerate(mob_headers, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border
        next_row += 1

        # ── Data rows ─────────────────────────────────────────────────────────
        for record in mobialert_rows:
            values = [
                record.get("CONFIGID"),
                record.get("CONFIGNAME"),
                record.get("EMAILMESSAGE"),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.alignment = left_align
                cell.border    = thin_border
            next_row += 1

        # ── Auto-width for the 3 MobiAlert columns (cols 1-3) ─────────────────
        # We re-scan only the rows we just wrote plus headers to avoid
        # fighting with the merged cells already in the sheet.
        mob_start_row = ws.max_row - len(mobialert_rows) - 1  # header + data
        for col_idx in range(1, 4):
            max_len = len(mob_headers[col_idx - 1])
            for row_idx in range(mob_start_row, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            col_letter = get_column_letter(col_idx)
            current_width = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(
                current_width, min(max_len + 2, 80)
            )

        wb.save(excel_path)
        logger.info(
            f"Appended {len(mobialert_rows)} MobiAlert row(s) to {excel_path}"
        )
        return excel_path

    except Exception as e:
        logger.error(f"Failed to append MobiAlert data to Excel: {e}")
        return None
